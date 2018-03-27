##   Organize our data using OPENPYXL
from openpyxl import load_workbook
import xlrd
import xlwt
import os
##Directory for infile
path = "C:/Users/bag019/Dropbox/JENSEN LAB DATA 2016/Compile/Data Excel Files"
infile_name = input('Enter infile name: ')
## Check sheet names
os.chdir(path)
wb = load_workbook(infile_name,data_only=True)
#list for 28 on
my_list =['WP AI', 'WP HX','PPT E','TC','HDL-C','TG','C3 (P1)','HP (P1)','E (P1)',
           'J #1 (P1)','C3 PPT (P1)','J  # 2 (P1)','AI(C3+) (P3)','AI(HP+) (P3)','AI(E+) (P3)',
           'AI(J+) #1 (P3)','E(AI+) (P3)','E(C3+) PPT (P3)','E(J+) #2 (P3)']


##run 25 on no a1 p1
##workded for 17 (check other new)['WP AI', 'WP HX','PPT E','TC','HDL-C','TG','C3 (P1)','HP (P1)','E (P1)',
##           'J #1 (P1)','C3 PPT (P1)','J  ppt (P1)','AI(C3+) (P3)','AI(HP+) (P3)','AI(E+) (P3)',
##           'AI(J+) #1 (P3)','E(AI+) (P3)','E(C3+) PPT (P3)','E(J+) #2 (P3)']

## runs 1-27 (not including 17) ['WP AI', 'WP HX','PPT E','TC','HDL-C','TG','AI (P1)','C3 (P1)','HP (P1)','E (P1)',
##           'J (P1)','C3 PPT (P1)','J PPT (P1)','AI(C3+) (P3)','AI(HP+) (P3)','AI(E+) (P3)',
##           'AI(J+) (P3)','E(AI+) (P3)','E(C3+) PPT (P3)','E(J+) PPT (P3)']


##run 17 needs the new data format then 28 on is new data format too
check = wb.get_sheet_names()
stop = 1
for item in my_list:
    if item not in check:
        stop = 0
        print(item)

# Function to compile sheets
def compiledata (data_from_sheet, data_to_txtfile):
    "This will copy data from a sheet and paste it to a sheet"
    os.chdir(path)
    wb = load_workbook(infile_name,data_only=True)
    ws_from = wb.get_sheet_by_name(data_from_sheet)


    #loop to copy data(this is the range of old template)
    #range of new template is (i 41, 76) (j 2,14)(j==13)
    #(1-27) for i in range(50,86):
    for i in range(41,77):
        for j in range(2,14):
           data_from = ws_from.cell(row = i, column = j).value
           os.chdir("""C:/Users/bag019/Dropbox/JENSEN LAB DATA 2016/Compile/txt files""")
           outfile = open(data_to_txtfile, 'a')
           outfile.write(str(data_from))
           outfile.write(' ')
           if j == 13:
               outfile.write('\n')


## sheet from and txt file to
if stop != 0:
    #SSE
    compiledata ('WP AI','WP AI.txt')
    compiledata ('WP HX','WP HX.txt')
    compiledata ('PPT E','PPT E.txt')
    #Lipids
    compiledata ('TC','TC.txt')
    compiledata ('HDL-C','HDL-C.txt')
    compiledata ('TG','TG.txt')
    #MSE (P1)
    compiledata ('C3 (P1)','C3 (P1).txt')
    compiledata ('HP (P1)','HP (P1).txt')
    compiledata ('E (P1)','E (P1).txt')
    compiledata ('J #1 (P1)','J (P1).txt')
    compiledata ('C3 PPT (P1)','C3 PPT (P1).txt')
    compiledata ('J  # 2 (P1)','J (P1) #2.txt')
    #MSE (P3) A1
    compiledata ('AI(C3+) (P3)','AI(C3+).txt')
    compiledata ('AI(HP+) (P3)','AI(HP+).txt')
    compiledata ('AI(E+) (P3)','AI(E+).txt')
    compiledata ('AI(J+) #1 (P3)','AI(J+).txt')
    #MSE (P3) E

    compiledata ('E(C3+) PPT (P3)','E(C3+) PPT.txt')
    compiledata ('E(J+) #2 (P3)','E(J+).txt')
    compiledata ('E(AI+) (P3)','PPT E(AI+).txt')


