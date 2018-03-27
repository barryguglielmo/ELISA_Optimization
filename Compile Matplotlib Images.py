# get jpg from file matplotlib images and put in an excel file
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

path1 = "C:/Users/bag019/Desktop/Matplotlib Images"
os.chdir(path1)
curve_images = []
cvs_images = []
sheet_names = ['WP AI', 'WP HX', 'PPT E', 'C3 (P1)', 'HP (P1)', 'E (P1)', 'J #1 (P1)', 'C3 PPT (P1)', 'AI(C3+) (P3)', 'AI(HP+) (P3)', 'AI(E+) (P3)', 'AI(J+) #1 (P3)', 'E(AI+) (P3)', 'E(C3+) PPT (P3)', 'E(J+) #2 (P3)']
wb1 = load_workbook('CHS_GRAPHS.xlsx')
wb2 = load_workbook('CHS_GRAPHS.xlsx')

for sheet in sheet_names:
    Anchor = 1
    for i in range(77, 135):
        if i < 100:
            try:
                img = Image("0"+str(i)+ ".xlsx curve " + sheet + '.png', size = (400,400))
                img.anchor(wb1[sheet]['A' +str(Anchor)])
                wb1[sheet].add_image(img)
                img2 = Image("0" + str(i)+ ".xlsx cvs " + sheet + '.png', size = (400,400))
                img2.anchor(wb1[sheet]['G' +str(Anchor)])
                wb1[sheet].add_image(img2)
                wb1[sheet].cell(row = Anchor, column = 20).value = i
                Anchor += 20
            except:
                wb1[sheet].cell(row = Anchor, column = 20).value = i
                Anchor += 20
        if i >= 100:
            try:
                img = Image(str(i)+".xlsx curve " + sheet + '.png', size = (400,400))
                img.anchor(wb2[sheet]['A' +str(Anchor)])
                wb2[sheet].add_image(img)
                img2 = Image(str(i)+".xlsx cvs " + sheet + '.png', size = (400,400))
                img2.anchor(wb2[sheet]['G' +str(Anchor)])
                wb2[sheet].add_image(img2)
                wb2[sheet].cell(row = Anchor, column = 20).value = i
                Anchor += 20
            except:
                wb2[sheet].cell(row = Anchor, column = 20).value = i
                Anchor += 20
print("Images done")


wb1.save("CHS_GRAPHS_[77-99].xlsx")
wb2.save("CHS_GRAPHS_[100-134].xlsx")
    


