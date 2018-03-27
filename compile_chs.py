#in the future make sure that all are moved over 1 and that run number appended
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
############################################
###################################
############
import os
import openpyxl
from openpyxl import load_workbook

os.chdir('D:/Jensen 2016-2017/CHS_Optimized_and_Final_ 077-134')

sheet_names = ['WP AI', 'WP HX', 'PPT E', 'C3 (P1)', 'HP (P1)', 'E (P1)', 'J #1 (P1)', 'C3 PPT (P1)', 'AI(C3+) (P3)', 'AI(HP+) (P3)', 'AI(E+) (P3)', 'AI(J+) #1 (P3)', 'E(AI+) (P3)', 'E(C3+) PPT (P3)', 'E(J+) #2 (P3)']
wb2 = load_workbook("Compiled Optimized CHS.xlsx")
files = []
for i in range(76,100):
    files.append("Optimized - 0" + str(i)+".xlsx")
for i in range(100,134):
    files.append("Optimized - " + str(i)+".xlsx")
for file in files:
    infile_name = file #input ("Enter infile name: ")
    wb = load_workbook(infile_name,data_only=True)
    for sheet in sheet_names:
        eof = 1
        try:
            ws_to = wb2.get_sheet_by_name(sheet)
            while (ws_to.cell(row = eof, column = 1).value) != None:
                eof +=1
            ws_from = wb.get_sheet_by_name(sheet)
            #columns
            a =[]
            b=[]
            c=[]
            d=[]
            e=[]
            f=[]
            g=[]
            h=[]
            ix=[]
            j=[]
            k=[]
            l=[]
            m=[]
            for i in range(41,77):
                a.append(ws_from.cell(row = i, column = 1).value)
                wb2[sheet].cell(row = i-41 + eof , column = 1).value = a[i-41]
            for i in range(41,77):
                b.append(ws_from.cell(row = i, column = 2).value)
                wb2[sheet].cell(row = i-41 + eof , column = 2).value = b[i-41]
            for i in range(41,77):
                c.append(ws_from.cell(row = i, column = 3).value)
                wb2[sheet].cell(row = i-41 + eof , column = 3).value = c[i-41]
            for i in range(41,77):
                d.append(ws_from.cell(row = i, column = 4).value)
                wb2[sheet].cell(row = i-41 + eof , column = 4).value = d[i-41]
            for i in range(41,77):
                e.append(ws_from.cell(row = i, column = 5).value)
                wb2[sheet].cell(row = i-41 + eof , column = 5).value = e[i-41]
            for i in range(41,77):
                f.append(ws_from.cell(row = i, column = 6).value)
                wb2[sheet].cell(row = i-41 + eof , column = 6).value = f[i-41]
            for i in range(41,77):
                g.append(ws_from.cell(row = i, column = 7).value)
                wb2[sheet].cell(row = i-41 + eof , column = 7).value = g[i-41]
            for i in range(41,77):
                h.append(ws_from.cell(row = i, column = 8).value)
                wb2[sheet].cell(row = i-41 + eof , column = 8).value = h[i-41]
            for i in range(41,77):
                ix.append(ws_from.cell(row = i, column = 9).value)
                wb2[sheet].cell(row = i-41 + eof , column = 9).value = ix[i-41]
            for i in range(41,77):
                j.append(ws_from.cell(row = i, column = 10).value)
                wb2[sheet].cell(row = i-41 + eof , column = 10).value = j[i-41]
            for i in range(41,77):
                k.append(ws_from.cell(row = i, column = 11).value)
                wb2[sheet].cell(row = i-41 + eof , column = 11).value = k[i-41]
            for i in range(41,77):
                l.append(ws_from.cell(row = i, column = 12).value)
                wb2[sheet].cell(row = i-41 + eof , column = 12).value = l[i-41]
            for i in range(41,77):
                m.append(ws_from.cell(row = i, column = 13).value)
                wb2[sheet].cell(row = i-41 + eof , column = 13).value = m[i-41]
        except:
            for i in range (0,36):
                for j in range (1,14):
                    wb2[sheet].cell(row = i-41 + eof , column = j).value = 0
            eof += 36
wb2.save("Compiled Optimized CHS (0).xlsx")
















