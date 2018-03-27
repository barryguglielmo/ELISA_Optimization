##import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import os
import plotly.plotly as py
from plotly.graph_objs import *
py.sign_in('barryguglielmo','WNMyBgUv3C0tHQzvfN33')


##os.chdir('C:/Users/bag019/Desktop') # work
os.chdir('C:/Users/Barry/Desktop')
wb = load_workbook('Database Fo Real.xlsx',data_only=True)
ws_from = wb.get_sheet_by_name('WP AI')
##data = 'Database Fo Real.csv'
##x = pd.read_csv(data, header = 0, usecols = ['S.ID'])
##y = pd.read_csv(data,header = 0, usecols = ['Average 1'], index_col = ['Average 1'])
##
y = []
x = []
for i in range(3,1837):
   data_from1 = ws_from.cell(row = i, column = 5).value
   y.append(data_from1)
for j in range(3,1837):
   data_from2 = ws_from.cell(row = j, column = 2).value
   x.append(data_from2)   
for thing in y:
   thing = str(thing)
   for letter in thing:
      if letter == "'":
         letter == ''

trace1 = {
  "x" : x,
  "y" : y, 
  "name": "WP AI", 
  "type": "bar"
}

data = Data([trace1])
fig = Figure(data=data)
plot_url = py.plot(fig)


##print (my_list_x)
##print (my_list_y)
##plt.bar(my_list_x,my_list_y,label = 'Yah doode')
##plt.show()

