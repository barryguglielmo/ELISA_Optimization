##   Organize our data using OPENPYXL
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
import xlsxwriter
from xlsxwriter import Workbook
import os
##Directory for infile
path = "C:/Users/bag019/Desktop/Data Excel Files"
## Check sheet names
os.chdir(path)

#list for 28 on
file_list=['028.xlsx','030.xlsx','031.xlsx']
my_list =['WP AI', 'WP HX','PPT E','C3 (P1)','HP (P1)','E (P1)',
           'J #1 (P1)','C3 PPT (P1)','J  # 2 (P1)','AI(C3+) (P3)','AI(HP+) (P3)','AI(E+) (P3)',
           'AI(J+) #1 (P3)','E(AI+) (P3)','E(C3+) PPT (P3)','E(J+) #2 (P3)']

wb2 = xlsxwriter.Workbook('Curves.xlsx')
for sheet in my_list:
    wb2.add_worksheet(sheet)

wb2 = load_workbook('Curves.xlsx')

##run 17 needs the new data format then 28 on is new data format too


    # Function to compile sheets
def cc (file, data_from_sheet, loci, dil):



    "This will compile curve data from a xlsx sheet and paste it to a txt sheet"
## To sheet
    
## Where the Poly comes from        
    ws_from = wb.get_sheet_by_name(data_from_sheet)
    a = ws_from.cell(row = 20, column = 20).value
    b = ws_from.cell(row = 20, column = 21).value
    c = ws_from.cell(row = 20, column = 22).value
    d = ws_from.cell(row = 20, column = 23).value
## Making a list of things to graph       
    x = 0
    x_list = []
    y_list = []
    a= float(a)
    b= float(b)
    c= float(c)
    d= float(d)
    org_loci = loci
    while x < 2.5:
        
        poly = (((a*(x**3))+(b*(x**2))+(c*(x))+(d))*dil)
        poly = str(poly)
        x = str(x)
##        print (x, poly)
        x_list.append(x)
        y_list.append(poly)
        x = float(x)+0.1
## Pasting graph data to Curves.xlsx
## This is being a pain in my ass
    ws_to = wb2.get_sheet_by_name(data_from_sheet)
    for that in x_list:
        ws_to.cell(row = int(loci), column = 1).value = that
        loci = int(loci) + 1
        
    loci = org_loci         
    for something in y_list:
        ws_to.cell(row = int(loci), column = 2).value = something
        loci = int(loci) + 1
    
    chart1 = ScatterChart()
# Configure the first series.
    chart1.title= file + ' ' + data_from_sheet
    chart1.style = 13
    chart1.x_axis.title = 'Absorbance'
    chart1.y_axis.title = 'mg/dL'
    mymin = str(org_loci)
    mymax = int(org_loci) + len(x_list)
    mymax = str(mymax)
    xvalues = Reference(ws_to, min_col = 1, min_row = mymin, max_row = mymax)
    yvalues = Reference(ws_to, min_col = 2, min_row = mymin, max_row = mymax)
    series = Series(yvalues, xvalues)
    chart1.series.append(series)

    ws_to.add_chart(chart1, 'C' + org_loci)

# Set an Excel chart style.
#chart1.set_style(11)       
##    ws_to.insert_chart('C'+loci, chart1)

loc = 1               
for file in file_list:
    wb = load_workbook(file, data_only = True)
    check = wb.get_sheet_names()
    stop = 1
    for item in my_list:
        if item not in check:
            stop = 0
            print(item)

    if stop != 0:
        for item in my_list:
            if item == 'WP AI':
                dil = 140000
            elif item == 'WP HX':
                dil = 10000
            elif item =='PPT E':
                dil = 2000
            elif item == 'C3 (P1)':
                dil = 10000
            elif item == 'HP (P1)':
                dil = 200000
            elif item == 'E (P1)':
                dil = 5000
            elif item == 'J #1 (P1)':
                dil = 5000
            elif item == 'C3 PPT (P1)':
                dil = 10000
            elif item == 'AI(C3+) (P3)':
                dil = 12000
            elif item == 'AI(HP+) (P3)':
                dil = 240000
            elif item == 'A1(E+) (P3)':
                dil = 6000
            elif item == 'A1(J+) #1 (P3)':
                dil = 6000
            elif item == 'E(AI+) (P3)':
                dil = 6000
            elif item == 'E(C3+) PPT (P3)':
                dil = 12000
            elif item =='E(J+) #2 (P3)':
                dil = 6000
            loc = str(loc)
            cc (file, item, loc,dil)
    loc = int(loc)+ 50
    wb2.save('Curves.xlsx')

