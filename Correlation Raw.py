## The challange here is having all the 16 sheets and all of their possible
## Itterations. How would I go about doing this?
## First it should be done in an excel sheet.
## How do I structure my lists?
##
import xlsxwriter
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Series, Reference
import os

os.chdir('C:/Users/Barry/Desktop/data graphs')
wb = xlsxwriter.Workbook('Graphs.xlsx')
wb2 = load_workbook("Database Fo Real.xlsx")


#use comb function


#### 1: WP AI
##AI_1 =['WP AI', 'WP HX']
##AI_2 = ['WP AI','PPT E']
##AI_3 = ['WP AI', 'TG',]
##AI_4=['WP AI', 'HDL-C']
##AI_5=['WP AI', 'TC']
##AI_6=['WP AI', 'AI (P1)']
##AI_7=['WP AI', 'C3 (P1)']
##AI_8=['WP AI', 'HP (P1)']
##AI_9=['WP AI', 'E (P1)']
##AI_10=['WP AI', 'J #1 (P1)']
##AI_11=['WP AI','C3 PPT (P1)']
##AI_12=['WP AI','J PPT (P1)']
##AI_13=['WP AI','AI(C3+) (P3)']
##AI_14=['WP AI','AI(HP+) (P3)']
##AI_15=['WP AI','AI(E+) (P3)']
##AI_16=['WP AI','AI(J+) #1 (P3)']
##AI_17=['WP AI','E(AI+) (P3)']
##AI_18=['WP AI','E(C3+) PPT (P3)']
##AI_19=['WP AI','E(J+) #2 (P3)']
####E(J)
##E(J)_1 =['E(J+) #2 (P3)', 'WP HX']
##E(J)_2 = ['E(J+) #2 (P3)','PPT E']
##E(J)_3 = ['E(J+) #2 (P3)', 'TG',]
##E(J)_4=['E(J+) #2 (P3)', 'HDL-C']
##E(J)_5=['E(J+) #2 (P3)', 'TC']
##E(J)_6=['E(J+) #2 (P3)', 'AI (P1)']
##E(J)_7=['E(J+) #2 (P3)', 'C3 (P1)']
##E(J)_8=['E(J+) #2 (P3)', 'HP (P1)']
##E(J)_9=['E(J+) #2 (P3)', 'E (P1)']
##E(J)_10=['E(J+) #2 (P3)', 'J #1 (P1)']
##E(J)_11=['E(J+) #2 (P3)','C3 PPT (P1)']
##E(J)_12=['E(J+) #2 (P3)','J PPT (P1)']
##E(J)_13=['E(J+) #2 (P3)','AI(C3+) (P3)']
##E(J)_14=['E(J+) #2 (P3)','AI(HP+) (P3)']
##E(J)_15=['E(J+) #2 (P3)','AI(E+) (P3)']
##E(J)_16=['E(J+) #2 (P3)','AI(J+) #1 (P3)']
##E(J)_17=['E(J+) #2 (P3)','E(AI+) (P3)']
##E(J)_18=['E(J+) #2 (P3)','E(C3+) PPT (P3)']
#### E(C3 PPT)
##E(C3)_1 =['E(C3+) PPT (P3)', 'WP HX']
##E(C3)_2 = ['E(C3+) PPT (P3)','PPT E']
##E(C3)_3 = ['E(C3+) PPT (P3)', 'TG',]
##E(C3)_4=['E(C3+) PPT (P3)', 'HDL-C']
##E(C3)_5=['E(C3+) PPT (P3)', 'TC']
##E(C3)_6=['E(C3+) PPT (P3)', 'AI (P1)']
##E(C3)_7=['E(C3+) PPT (P3)', 'C3 (P1)']
##E(C3)_8=['E(C3+) PPT (P3)', 'HP (P1)']
##E(C3)_9=['E(C3+) PPT (P3)', 'E (P1)']
##E(C3)_10=['E(C3+) PPT (P3)', 'J #1 (P1)']
##E(C3)_11=['E(C3+) PPT (P3)','C3 PPT (P1)']
##E(C3)_12=['E(C3+) PPT (P3)','J PPT (P1)']
##E(C3)_13=['E(C3+) PPT (P3)','AI(C3+) (P3)']
##E(C3)_14=['E(C3+) PPT (P3)','AI(HP+) (P3)']
##E(C3)_15=['E(C3+) PPT (P3)','AI(E+) (P3)']
##E(C3)_16=['E(C3+) PPT (P3)','AI(J+) #1 (P3)']
##E(C3)_17=['E(C3+) PPT (P3)','E(AI+) (P3)']
#### E (AI)
##E(AI)_1 =['E(AI+) (P3)', 'WP HX']
##E(AI)_2 = ['E(AI+) (P3)','PPT E']
##E(AI)_3 = ['E(AI+) (P3)', 'TG',]
##E(AI)_4=['E(AI+) (P3)', 'HDL-C']
##E(AI)_5=['E(AI+) (P3)', 'TC']
##E(AI)_6=['E(AI+) (P3)', 'AI (P1)']
##E(AI)_7=['E(AI+) (P3)', 'C3 (P1)']
##E(AI)_8=['E(AI+) (P3)', 'HP (P1)']
##E(AI)_9=['E(AI+) (P3)', 'E (P1)']
##E(AI)_10=['E(AI+) (P3)', 'J #1 (P1)']
##E(AI)_11=['E(AI+) (P3)','C3 PPT (P1)']
##E(AI)_12=['E(AI+) (P3)','J PPT (P1)']
##E(AI)_13=['E(AI+) (P3)','AI(C3+) (P3)']
##E(AI)_14=['E(AI+) (P3)','AI(HP+) (P3)']
##E(AI)_15=['E(AI+) (P3)','AI(E+) (P3)']
##E(AI)_16=['E(AI+) (P3)','AI(J+) #1 (P3)']
#### AI (J)
##AI(J)_1 =['AI(J+) #1 (P3)', 'WP HX']
##AI(J)_2 = ['AI(J+) #1 (P3)','PPT E']
##AI(J)_3 = ['AI(J+) #1 (P3)', 'TG',]
##AI(J)_4=['AI(J+) #1 (P3)', 'HDL-C']
##AI(J)_5=['AI(J+) #1 (P3)', 'TC']
##AI(J)_6=['AI(J+) #1 (P3)', 'AI (P1)']
##AI(J)_7=['AI(J+) #1 (P3)', 'C3 (P1)']
##AI(J)_8=['AI(J+) #1 (P3)', 'HP (P1)']
##AI(J)_9=['AI(J+) #1 (P3)', 'E (P1)']
##AI(J)_10=['AI(J+) #1 (P3)', 'J #1 (P1)']
##AI(J)_11=['AI(J+) #1 (P3)','C3 PPT (P1)']
##AI(J)_12=['AI(J+) #1 (P3)','J PPT (P1)']
##AI(J)_13=['AI(J+) #1 (P3)','AI(C3+) (P3)']
##AI(J)_14=['AI(J+) #1 (P3)','AI(HP+) (P3)']
##AI(J)_15=['AI(J+) #1 (P3)','AI(E+) (P3)']




mylist = [AI_1, AI_2,AI_3,AI_4,AI_5,AI_6,AI_7,AI_8,AI_9,AI_10,AI_11,AI_12,AI_13,AI_14,AI_15,AI_16,AI_17,AI_18,AI_19]


def corrilation (couple):
    "Input the POI for corrilation"
    counter = 0
    for item in couple:
        if counter == 0:
            x = item
            counter += 1
        elif counter == 1:
            y = item
    print (x,y)
    xname= str(x)
    yname= str(y)
    worksheet = wb.add_worksheet(xname +' vs ' + yname)
##    paste data into appropriate sheet for refference
    xlist = []
    ylist = []
    ws_from1 = wb2.get_sheet_by_name(x)
    for i in range(3, 1837):
        data_from1 = ws_from1.cell(row = i, column = 11).value
        ylist.append(data_from1)
    ws_from2 = wb2.get_sheet_by_name(y)
    for j in range(3, 1837):
        data_from2 = ws_from2.cell(row = j, column = 11).value
        xlist.append(data_from2)
    
## Put this data into the graphs xlx
    
    worksheet.write_column('B2', xlist)
    worksheet.write_column('C2', ylist)

####    geter done
    chart1 = wb.add_chart({'type': 'scatter'})
    # Configure the first series.
    chart1.add_series({
        'name': xname + ' vs ' + yname,
        'categories': ('=' + "'" + xname +' vs ' + yname + "'" + '!' + '$B$2:$B$1837'),
        'values': ('=' + "'" + xname +' vs ' + yname + "'"  + '!' + '$C$2:$C$1837'),
        'trendline': {'type': 'linear'}
    })

    # Add a chart title and some axis labels.
    chart1.set_title ({'name': xname + ' vs ' + yname})
    chart1.set_x_axis({'name': xname})
    chart1.set_y_axis({'name': yname})

    # Set an Excel chart style.
    chart1.set_style(11)

# Insert the chart into the worksheet (with an offset).
    
    worksheet.insert_chart('A1', chart1)


for couple in mylist:
    corrilation (couple)

wb.close()
##    


