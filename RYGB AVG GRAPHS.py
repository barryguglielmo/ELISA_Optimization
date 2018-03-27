##import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Series, Reference
##import plotly.plotly as py
import os
##py.sign_in('barryguglielmo','WNMyBgUv3C0tHQzvfN33')
os.chdir('C:/Users/bag019/Dropbox/JENSEN LAB DATA 2016/Database Excel Sheets')
## should i really do data only?
wb = load_workbook('RYGB AVG.xlsx')
#wb2 = load_workbook('RYGB AVG 2.xlsx')
##raw data
tabs = ['WP AI', 'WP HX','PPT E','TC','HDL-C','TG','C3 (P1)','HP (P1)','E (P1)',
           'J #1 (P1)','C3 PPT (P1)','AI(C3+) (P3)','AI(HP+) (P3)','AI(E+) (P3)',
           'AI(J+) #1 (P3)','E(AI+) (P3)','E(C3+) PPT (P3)','E(J+) #2 (P3)']
def RYGB (in_sheet, col, paste, color):
   "Make RYGB chart for each in_sheet"

   ws_from = wb.get_sheet_by_name(in_sheet)
   chart1 = BarChart()
   chart1.type = "col"
   chart1.style = 6
   chart1.title = in_sheet + ' ' + color
   chart1.y_axis.title = 'mg/dL'
   chart1.x_axis.title = 'Run #'
   data = Reference(ws_from, min_col= col, min_row=3, max_row= 77)
   chart1.add_data(data)
   chart1.shape = 4
   my_graph = wb.get_sheet_by_name(in_sheet)
   my_graph.add_chart(chart1, 'J'+ str(paste))
   wb.save("TRY RYGB.xlsx")

   
   
for tab in tabs:
   RYGB(tab,2,2, "RED")
   RYGB(tab,4,22, "Yellow")
   RYGB(tab,6,42,"Green")
   RYGB(tab,8,62, "Blue")
   


wb.save('AVG GRAPHS.xlsx') 
## need to know how to play with x axis
