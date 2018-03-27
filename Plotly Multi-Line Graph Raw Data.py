import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import os
import plotly.plotly as py
from plotly.graph_objs import *
py.sign_in('barryguglielmo','WNMyBgUv3C0tHQzvfN33')


os.chdir('C:/Users/bag019/Desktop') # work
##os.chdir('C:/Users/Barry/Desktop')
wb = load_workbook('Database For Editing.xlsx',data_only=True)
mylist = ['WP AI', 'WP HX', 'PPT E', 'TG', 'HDL-C', 'TC', 'C3 (P1)',
          'HP (P1)', 'E (P1)', 'J #1 (P1)', 'C3 PPT (P1)',  'J PPT (P1)',
          'AI(C3+) (P3)', 'AI(HP+) (P3)', 'AI(E+) (P3)',
          'E(AI+) (P3)', 'E(C3+) PPT (P3)', 'E(J+) #2 (P3)']
##'J  # 2 (P1)',
##'AI(J+) #1 (P3)'
##'AI (P1)',
count = 1
tracers = []
a = 0
b = 0
c = 0
for sheet in mylist:
    trace = 'trace' + str(count)
    
    
    ws_from = wb.get_sheet_by_name(sheet)
    y = []
    x = []

    for i in range(3,1795):
       data_from1 = ws_from.cell(row = i, column = 2).value
       y.append(data_from1)
    for j in range(3,1795):
       data_from2 = ws_from.cell(row = j, column = 11).value
       x.append(data_from2)   
    for thing in y:
       thing = str(thing)
       for letter in thing:
          if letter == "'":
             letter == ''
    trace = { "x": y, "y": x, "line": {"color": "rgb"+str(a)+','+str(b)+','+str(c), "width": 1}, 
    "name": sheet, "opacity": 0.49, "type": "scatter", "uid": "a14820", 
    "xsrc": "maartenzam:231:JMZVJBWNUQKK2EXMN3M0N9FTMVQVAVTA", 
    "ysrc": "maartenzam:231:QF5ASZ6I2M18JEHCASB3NPONZ7T07J0J"}
    tracers.append(trace)
    count += 1
    a += 10
    b += 10
    c += 10
    
data = Data(tracers)
##layout = {"annotations": [{"x": 1.40192331531e+12,"y": 28.4293867551, 
##      "font": {"color": "rgb(214, 39, 40)", "size": 16}, "showarrow": False, "text": "Griekenland","textangle": 0}, 
##    {"x": 1.4035873857e+12, "y": 22.8150025602, "font": {"color": "rgb(255, 127, 14)", "size": 16}, "showarrow": False, "text": "Spanje"}, 
##    {"x": 1.40881732123e+12, "y": 11.479109063, "font": {"size": 16}, "showarrow": False, "text": "EU28"}], 
##  "autosize": True, 
##  "height": 614, 
##  "hovermode": "closest", 
##  "paper_bgcolor": "rgb(255, 247, 234)", 
##  "plot_bgcolor": "rgb(255, 247, 234)", 
##  "showlegend": False, 
##  "title": "Werkloosheid in de EU: torenhoog in Griekenland en Spanje", 
##  "titlefont": {"color": "rgb(102, 102, 102)", "family": "Droid Serif, serif", "size": 24}, 
##  "width": 1166, 
##  "xaxis": {"autorange": True, "range": [946681200000, 1.4249825765e+12], "rangemode": "tozero", "showgrid": False, "showline": True, 
##            "tickfont": { "color": "rgb(102, 102, 102)",  "family": "Droid Serif, serif", "size": 14}, 
##  "titlefont": { "color": "rgb(204, 204, 204)", "family": "Droid Serif, serif", "size": 16}, "type": "date"}, 
##  "yaxis": {"autorange": True, "range": [0, 300], "tickfont": {"color": "rgb(102, 102, 102)", "family": "Droid Serif, serif", "size": 14}, 
##  "title": "Werkloosheidspercentage", 
##  "titlefont": {"color": "rgb(102, 102, 102)", "family": "Droid Sans, sans-serif", "size": 16}, 
##    "type": "linear"
##  }
##}
fig = Figure(data=data)
plot_url = py.plot(fig)   
