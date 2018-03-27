import os
import openpyxl
from openpyxl import load_workbook

os.chdir('C:/Users/bag019/Desktop/Data Corrected Files')

sheet_names = ['WP AI', 'WP HX', 'PPT E', 'C3 (P1)', 'HP (P1)', 'E (P1)', 'J #1 (P1)', 'C3 PPT (P1)', 'AI(C3+) (P3)', 'AI(HP+) (P3)', 'AI(E+) (P3)', 'AI(J+) #1 (P3)', 'E(AI+) (P3)', 'E(C3+) PPT (P3)', 'E(J+) #2 (P3)']
wb2 = load_workbook("Compiled CHS.xlsx")
files = []
for i in range(77,100):
    files.append("Optimized - 0" + str(i)+".xlsx")
for i in range(100,120):
    files.append("Optimized - " + str(i)+".xlsx")
print(files)
for file in files:
    infile_name = file #input ("Enter infile name: ")
    wb = load_workbook(infile_name,data_only=True)
    for sheet in sheet_names:
        eof = 2
        ws_to = wb2.get_sheet_by_name(sheet)
        while (ws_to.cell(row = eof, column = 32).value) != None:
            eof += 36
        ws_from = wb.get_sheet_by_name(sheet)
        #columns
        a =[]
        b=[]
        c=[]
        d=[]
        e=[]
        f=[]
        g=[]
        h=[]
        ix=[]
        j=[]
        k=[]
        l=[]
        m=[]
        for i in range(77,81):
            a.append(ws_from.cell(row = i, column = 1).value)
            wb2[sheet].cell(row = i-77 + eof , column = 32).value = a[i-77]
        for i in range(77,81):
            b.append(ws_from.cell(row = i, column = 2).value)
            wb2[sheet].cell(row = i-77 + eof , column = 33).value = b[i-77]
        for i in range(77,81):
            c.append(ws_from.cell(row = i, column = 3).value)
            wb2[sheet].cell(row = i-77 + eof , column = 34).value = c[i-77]
        for i in range(77,81):
            d.append(ws_from.cell(row = i, column = 4).value)
            wb2[sheet].cell(row = i-77 + eof , column = 35).value = d[i-77]
        for i in range(77,81):
            e.append(ws_from.cell(row = i, column = 5).value)
            wb2[sheet].cell(row = i-77 + eof , column = 36).value = e[i-77]
        for i in range(77,81):
            f.append(ws_from.cell(row = i, column = 6).value)
            wb2[sheet].cell(row = i-77 + eof , column = 37).value = f[i-77]
        for i in range(77,81):
            g.append(ws_from.cell(row = i, column = 7).value)
            wb2[sheet].cell(row = i-77 + eof , column = 38).value = g[i-77]
        for i in range(77,81):
            h.append(ws_from.cell(row = i, column = 8).value)
            wb2[sheet].cell(row = i-77 + eof , column = 39).value = h[i-77]
        for i in range(77,81):
            ix.append(ws_from.cell(row = i, column = 9).value)
            wb2[sheet].cell(row = i-77 + eof , column = 40).value = ix[i-77]
        for i in range(77,81):
            j.append(ws_from.cell(row = i, column = 10).value)
            wb2[sheet].cell(row = i-77 + eof , column = 77).value = j[i-77]
        for i in range(77,81):
            k.append(ws_from.cell(row = i, column = 11).value)
            wb2[sheet].cell(row = i-77 + eof , column = 42).value = k[i-77]
        for i in range(77,81):
            l.append(ws_from.cell(row = i, column = 12).value)
            wb2[sheet].cell(row = i-77 + eof , column = 43).value = l[i-77]
        for i in range(77,81):
            m.append(ws_from.cell(row = i, column = 13).value)
            wb2[sheet].cell(row = i-77 + eof , column = 44).value = m[i-77]
wb2.save("Compiled CHS.xlsx")

