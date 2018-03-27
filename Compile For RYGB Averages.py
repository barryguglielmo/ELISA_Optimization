##   Organize our data using OPENPYXL
from openpyxl import load_workbook
import xlrd
import xlwt
import os
##Directory for infile
path = "S:/Nutrition/Sacks Lab Data/2015.07 - Jensen - Lipoproteins & Brain/Data/Compiled Data"
infile_name = '4-20-17-Database For Run Analysis.xlsx'
os.chdir(path)
wb = load_workbook(infile_name,data_only=True)
#list of tabs
my_list =['WP AI', 'WP HX','PPT E','TC','HDL-C','TG','C3 (P1)','HP (P1)','E (P1)',
           'J #1 (P1)','C3 PPT (P1)','AI(C3+) (P3)','AI(HP+) (P3)','AI(E+) (P3)',
           'AI(J+) #1 (P3)','E(AI+) (P3)','E(C3+) PPT (P3)','E(J+) #2 (P3)']

# Function to compile sheets
def compileRYGB (data_from_sheet):
    "This will Compile the RYGB values"
    os.chdir(path)
    wb = load_workbook(infile_name,data_only=True)
    ws_from = wb.get_sheet_by_name(data_from_sheet)
    #Lists
    R = []
    Y = []
    G = []
    B = []
    #Cell to start
    i = 3
    j = 4
    k = 5
    l = 6
    #Loop to get R Values
    while i < 2671:
        data_from = ws_from.cell(row = i, column = 26).value
        R.append(data_from)
        i += 36
    #Loop to get Y Values
    while j < 2671:
        data_from = ws_from.cell(row = j, column = 26).value
        Y.append(data_from)
        j += 36
    #Loop to get G Values
    while k < 2671:
        data_from = ws_from.cell(row = k, column = 26).value
        G.append(data_from)
        k += 36
    #Loop to get B Values
    while l < 2671:
        data_from = ws_from.cell(row = l, column = 26).value
        B.append(data_from)
        l += 36

    os.chdir("""C:/Users/bag019/Dropbox/JENSEN LAB DATA 2016/Compile/RYGB/RYGB Avgs""")
    data_to_txtfile = data_from_sheet + '.txt'
    outfile = open(data_to_txtfile, 'a')
    outfile.write(str(R)+'\n' + '\n'+str(Y)+'\n' + '\n'+str(G)+'\n' + '\n'+str(B))


           


## sheet from and txt file to
for item in my_list:
    compileRYGB (item)
