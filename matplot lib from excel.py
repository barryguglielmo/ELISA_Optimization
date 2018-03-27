##import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import os
import plotly.plotly as py
##py.sign_in('barryguglielmo','WNMyBgUv3C0tHQzvfN33')


##os.chdir('C:/Users/bag019/Desktop') # work
os.chdir('C:/Users/Barry/Desktop')
wb = load_workbook('Database Fo Real.xlsx',data_only=True)
##raw data
def rawbar (in_sheet):
   "What raw data sheet to chart"
   ws_from = wb.get_sheet_by_name('WP AI')
   y = []
   z= []
   for i in range(3,1837):
      data_from1 = ws_from.cell(row = i, column = 11).value
      y.append(data_from1)
   for j in range(3,1837):
      data_from2 = ws_from.cell(row = j, column = 2).value
      z.append(data_from2)
   x = len(y)
   x_list = []
   b=0
   while b < x:
      b = b+1
      x_list.append(b)
   LABELS = z     
   plt.bar(x_list, y, align='center')
   plt.xticks(x_list, LABELS)
   plt.title(in_sheet + " Raw Data")
   plt.ylabel('Absorbance')
   plt.xlabel('Sample I.D.')
   plt.savefig('C:/Users/Barry/Desktop/data graphs/' + in_sheet + 'Raw Data.png')
##control data 
def controlbar (in_sheet):
   "What raw data sheet to chart"
   ws_from = wb.get_sheet_by_name('WP AI')
   y = []
   z= []
   for i in range(3,1837):
      data_from1 = ws_from.cell(row = i, column = 39).value
      y.append(data_from1)
   for j in range(3,1837):
      data_from2 = ws_from.cell(row = j, column = 2).value
      z.append(data_from2)
   x = len(y)
   x_list = []
   b=0
   while b < x:
      b = b+1
      x_list.append(b)
   LABELS = z     
   plt.bar(x_list, y, align='center')
   plt.xticks(x_list, LABELS)
   title = plt.title(in_sheet + " Control Data")
   plt.ylabel('Absorbance')
   plt.xlabel('Sample I.D.')
   plt.savefig('C:/Users/Barry/Desktop/data graphs/' + in_sheet + 'Control Data.png')
rawbar("WP AI")
##rawbar("WP HX")
controlbar("WP AI")
