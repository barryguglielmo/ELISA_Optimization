import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
def corrilation (sheet, x_vals, y_vals, best_m, loc, out_f):
    plt.clf()
    plt.scatter(x_vals,x_vals, color = 'b', marker = 'o', s = 50)
    plt.plot([0,max(x_vals)*2], [0, (max(x_vals)*2)*best_m],color = 'r',lw=1, label = "Slope = " + str(best_m), )
    plt.xlabel('Hand Done Data')
    plt.ylabel('Machine Done Data')
    plt.title(sheet + " Hand vs Machine")
    plt.xlim([0,max(x_vals)*2])
    plt.ylim([0,max(y_vals)*2])
    plt.legend()
    os.chdir("C:/Users/bag019/Desktop/Matplotlib Images")
    plt.savefig('050-' + sheet + ".png")
    img = Image( '050-' + sheet + ".png", size = (400,400))
    img.anchor(out_f['Regression']['A' + str(loc)])
    out_f['Regression'].add_image(img)
    plt.close()

first_row_of_data = 35
avg_column = 13
#MY FILE
os.chdir("C:/Users/bag019/Desktop/Data Corrected Files")
wb_op = load_workbook('Optimized - NS-042.xlsx', data_only = True)
out_f = load_workbook('Hand vs Machine.xlsx', data_only = True)
#FILE TO COMPARE TO
os.chdir("S:/Nutrition/Sacks Lab Data/2013.10 - HDL Proteins and CHD/DATA/RUNS")
wb_hand = load_workbook('2017.06.20 - MG NS - 042 - Corrected.xlsx', data_only = True)
sheet_names = ['WP AI', 'A1AT SSE', 'A2M SSE', 'FBG SSE', 'PLMG SSE', 'A2 P1', 'A4 P1', 'C1 P1', 'C2 P1', 'C3 P1', 'CoC3 P1', 'CP P1', 'E P1', 'HP P1', 'J P1', 'L1 P1', 'PON1 P1', 'E(A1)', 'A1AT P3', 'A2 P2', 'A2 P3', 'A4 P3', 'A2M P3', 'C1 P3', 'C2 P3', 'C3 P3', 'CoC3 P3', 'CP P3', 'E P3', 'FBG P3', 'HP P3', 'J P3', 'L1 P3', 'PLMG P3', 'PON1 P3']
loc = 1
for sheet in sheet_names:
    
    average_hand = []
    average_op = []
    for i in range(first_row_of_data, first_row_of_data + 36):
        average_hand.append(wb_hand[sheet].cell(row = i, column = avg_column).value)
        average_op.append(wb_op[sheet].cell(row = i, column = avg_column).value)
        #x = obtained amounts of RYGB, y = expected amounts of RYGB
    best_dist = 5
    m = 0
    best_m = 0                     
        ##sum of least squares to find the best slope ((y - mx)= d)
    while m < 2.5:
        sum_dsqrd = 0
        for i in range (0,36):
            x = average_hand[i]
            y = average_op[i]
            distance_sqrd = (y - (m*x))**2
            sum_dsqrd += distance_sqrd
            if sum_dsqrd < best_dist:
                best_dist = sum_dsqrd
                best_m = m
                m += 0.0001
            else:
                m += 0.0001
    print(sheet + " : " + str (best_m))
    corrilation(sheet, average_hand, average_op, best_m, loc, out_f)
    loc += 20
    os.chdir("C:/Users/bag019/Desktop/Data Corrected Files")
    out_f.save("050 - Hand vs Machine.xlsx")
