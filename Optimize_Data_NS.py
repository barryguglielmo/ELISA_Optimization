#The purpose of this program is to analyze data in duplicate for 96 well plates for ELISA assays and Lipids.
#If this program was written for use by teams under the direction of Dr. Jensen and Dr. Sacks.
#any other use of this program must have express permission from Barry Guglielmo.

#This program was written for Python 3.6
#pip install the following (matplotlib, openpyxl, image, pillow, numpy, )
################################################
#!!!!!!!!!!!!!!!TO DO!!!!!!!!!!!!!!!!!!!!!!!!!!#
################################################

## colors = http://www.color-hex.com/color/1072ba

#Do a comparison of hand optimized data

#Finish the Note step for obt being too high or too low compared to expected

#Better Flagging for Notes or Bad Data
    #cv's or outliers

#defend against empty sheets

#Think about making the places you expect data into variables

########CODE STARTS HERE#########
#########  HAVE FUN   ###########
#################################

#imports
import openpyxl
from openpyxl import load_workbook
from openpyxl import formatting, styles
from openpyxl.drawing.image import Image
from openpyxl.chart import ScatterChart, BarChart, Reference, Series

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

import numpy
import statistics

import os

#Graphing defs
def CURVE_VS_SAMP(sheet, my_x, samples ):
    #my_x = CURVE ABS, samples = SAMP ABS
    plt.clf()
    plt.scatter(samples,samples, label = "Curve Abs", color = 'r', marker = 'o', s = 75)
    plt.scatter(my_x,my_x, label = "Curve Abs", color = 'b', marker = 'o', s = 200)
    plt.xlabel('Absorbance')
    plt.ylabel('Absorbance')
    plt.title(sheet + " Curve vs Samples")
    plt.legend()
    os.chdir("C:/Users/bag019/Desktop/Matplotlib Images")
    plt.savefig(infile_name + ' cvs ' + sheet + ".png")
    img = Image(infile_name + ' cvs ' + sheet + ".png", size = (300,300))
    img.anchor(wb[sheet]['N5'])
    wb[sheet].add_image(img)
    plt.close()
def CURVE (sheet, my_x, my_y, best_poly):
    #my_x = curve abs, my_y = constant
    plt.clf()
    plt.scatter(my_x,my_y, label = "curve", color = 'b', marker = 'o', s = 200)
    plt.xlabel('Absorbance')
    plt.ylabel('Constant')
    plt.title(sheet + " Curve")
    #plot the poly line
    poly_xs = []
    i = 0
    while i < max(my_x)*1.25:
        poly_xs.append(i)
        i += 0.01
    poly_ys = []
    for item in poly_xs:
        poly_ys.append((numpy.polyval(best_poly, item)))
    plt.plot(poly_xs, poly_ys, color = 'b', lw = 1, label = "Poly of curve")
    plt.xlim([0,max(my_x)*1.25])
    plt.ylim([0,max(my_y)*1.5])
    Rsqrd = mpatches.Patch(color= 'green', label = 'R = ' + str(best_r) + "\npoly = " + str(best_poly))
    plt.legend(handles= [Rsqrd])
    os.chdir("C:/Users/bag019/Desktop/Matplotlib Images")
    plt.savefig(infile_name + ' curve ' + sheet + ".png")
    img2 = Image(infile_name + ' curve ' + sheet + ".png", size = (300,300))
    img2.anchor(wb[sheet]['N17'])
    wb[sheet].add_image(img2)
    plt.close()
    #exp vs obt rygb
    #~~~~good~~~~#
def EXP_VS_OBT (sheet, obtCor, expCor):
    #obtCor = obtained amounts of RYGB, expCor = expected amounts of RYGB
        plt.clf()
        plt.scatter(obtCor,expCor, color = 'b', marker = 'o', s = 50)
        plt.plot([0,max(obtCor)*2], [0, (max(obtCor)*2)*best_m],color = 'r',lw=1, label = "Slope = " + str(best_m), )
        plt.xlabel('Obtained')
        plt.ylabel('Expected')
        plt.title(sheet + " Exp vs Obt")
        plt.xlim([0,max(obtCor)*2])
        plt.ylim([0,max(expCor)*2])
        plt.legend()
        os.chdir("C:/Users/bag019/Desktop/Matplotlib Images")
        plt.savefig(infile_name + ' expvsobt ' + sheet + ".png")
        img3 = Image(infile_name + ' expvsobt ' + sheet + ".png", size = (400,400))
        img3.anchor(wb[sheet]['N29'])
        wb[sheet].add_image(img3)
        plt.close()
    #bargraph of exp and obt and corrected
def BAR_EXP_VS_OBT (sheet, obtCor, expCor , cor):
    #obtCor = obtained amounts of RYGB
    #expCor = expected amounts of RYGB
    #cor = corrected amounts of RYGB
    plt.clf()
    #number of clusters
    N = 4
    ind = numpy.arange(N)
    multiple_bars = plt.figure()
    #labels
    x = ['R','Y','G','B']
    y = obtCor
    z= expCor
    k= cor
    #obt
    plt.bar(ind, y,width=0.2,color='b',align='center', label = "Obtained")
    #exp
    plt.bar(ind+0.2, z,width=0.2,color='g',align='center', label ="Expected")
    #cor
    plt.bar(ind+0.2*2, k,width=0.2,color='r',align='center', label = "Corrected")
    #show color and meaning
    plt.legend()
    #x labels RYGB
    plt.xticks([0.3,1.3,2.3,3.3],x)
    os.chdir("C:/Users/bag019/Desktop/Matplotlib Images")
    plt.savefig(infile_name + ' c ' + sheet + ".png")
    img4 = Image(infile_name + ' c ' + sheet + ".png", size = (400,400))
    img4.anchor(wb[sheet]['A17'])
    wb[sheet].add_image(img4)
    plt.close()
def AVERAGES (sheet, samp_avg):
    #samp_avg = sample averages
    plt.clf()
    j = 1
    objects = []
    for i in range(0,40):
        objects.append(j)
        j +=1
    y_pos = numpy.arange(len(objects))                 
    plt.bar(y_pos, samp_avg, align='center', alpha=0.5)
    plt.xticks(y_pos, objects)
    plt.ylabel('mg/dl')
    plt.title('Sample Averages '+ sheet)
    os.chdir("C:/Users/bag019/Desktop/Matplotlib Images")
    plt.savefig(infile_name + ' avg ' + sheet + ".png")
    img5 = Image(infile_name + ' avg ' + sheet + ".png", size = (400,400))
    img5.anchor(wb[sheet]['H17'])
    wb[sheet].add_image(img5)
    plt.close

#START OF THE PROJECT
#To alter this for different layouts
#curve values (dil, y = const, x = abs)
row_of_curve = 18
col_of_curve_x = 24
col_of_curve_y = 23
col_of_curve_dil = 22
#dilution of samp
row_dil_samp = 2
col_dil_samp = 1
#row samples start
row_samp_start = 35
#column samples start
col_samp_start = 3
#Expected RYGB row and column
row_exp = 2
col_exp = 22
##Good_to_Go = True
#declare paths
path1 = "C:/Users/bag019/Desktop/Data Excel Files"
path2 = "C:/Users/bag019/Desktop/Data Corrected Files"

##if Good_to_Go == True:
files = ['NS-051.xlsx']

##for i in range (77,83):
##    files.append('0'+str(i)+'.xlsx')
##for i in range (85,100):
##    files.append('0'+str(i)+'.xlsx')
##for i in range (100,121):
##    files.append(str(i)+'.xlsx')
for file in files:
    print (file)
    infile_name = file #input('Enter infile name: ')
    os.chdir(path2)
    outfile = open("Notes", "a")
    outfile.write("File: " + infile_name + "\n\n")
    os.chdir(path1)
    #open up the excel workbook
    wb = load_workbook(infile_name,data_only=True)
    #list of the sheet names
    sheet_names = ['WP AI', 'A1AT SSE', 'A2M SSE', 'FBG SSE', 'PLMG SSE', 'A2 P1', 'A4 P1', 'C1 P1', 'C2 P1', 'C3 P1', 'CoC3 P1', 'CP P1', 'E P1', 'HP P1', 'J P1', 'L1 P1', 'PON1 P1', 'E(A1)', 'A1AT P3', 'A2 P2', 'A2 P3', 'A4 P3', 'A2M P3', 'C1 P3', 'C2 P3', 'C3 P3', 'CoC3 P3', 'CP P3', 'E P3', 'FBG P3', 'HP P3', 'J P3', 'L1 P3', 'PLMG P3', 'PON1 P3']
    #Lipids
    lipids = [ 'HDL-C']#, 'TG', 'TC']
    #to do for each sheet
    for sheet in sheet_names:
        print(sheet)
        ws_from = wb.get_sheet_by_name(sheet)
        go = True
        
        #Y value lists curve
        curve_1 = []
        curve_2 = []
        
        #x value lists curve
        curve_yval_1 = []
        curve_yval_2 = []

        
        #sample abs
        sample_1 = []
        sample_2 = []
        
        #sample amt (to be determinded using best poly)
        samp_amt1 = []
        samp_amt2 = []

        #sample avg
        samp_avg = []
        
        #correction
        corExp = []
        corObt = []
        
        #dilution of curve
        dil = []
        
        #dilution of samp
        samp_dil = ws_from.cell(row = row_dil_samp, column = col_dil_samp).value

        #correction obt and exp
        expCor = []
        obtCor = []

        #cv's
        cv = []

        #standard deviation
        stdv = []

        
    #####read the appropriate cells into each list
        #load x values
        #~~~~~~~~~~~~good~~~~~~~~~~~~~~~#
        #load curve_1 (x values)
        for i in range(row_of_curve , row_of_curve + 8):
            data_from = ws_from.cell(row = i, column = col_of_curve_x).value
            curve_1.append(data_from)
        #load curve_2 (x values)
        for i in range(row_of_curve + 8, row_of_curve + 16):
            data_from = ws_from.cell(row = i, column = col_of_curve_x).value
            curve_2.append(data_from)


        #load curve_yval_1 and 2 (y values) constants and dilutions
        #~~~~~~~~~~~good~~~~~~~~~~~~~~~~#
        for i in range(row_of_curve, row_of_curve + 8):
            data_from = ws_from.cell(row = i, column = col_of_curve_y).value
            curve_yval_1.append(data_from)
            curve_yval_2.append(data_from)
            if i == row_of_curve + 7:
                dil.append(0)
            else:
                data_from = ws_from.cell(row = i, column = col_of_curve_dil).value
                dil.append(data_from)



        
        #load sample abs (go a step farther and hilight samples ovrflw
        for i in range(row_samp_start, row_samp_start + 40):
            data_from = ws_from.cell(row = i, column = 3).value
            sample_1.append(data_from)
            data_from = ws_from.cell(row = i, column = 7).value
            sample_2.append(data_from)
        for i in range(0,40):
            if sample_1[i] == 'OVRFLW':
                if sample_2[i] == 'OVRFLW':
                    sample_1[i] = 4
                    sample_2[i] = 4
                    outfile.write(sheet+" Sample Overflow: Sample Number: " + str(i + 1) + " absorbance replaced with '4'\n")
                else:
                    sample_1[i] = sample_2[i]
                    outfile.write(sheet+" Sample Overflow: Sample Number: " + str(i + 1) + " absorbance replaced with second measured absorbance " + str(sample_2[i]) + "\n")

        for i in range(0,40):
            if sample_2[i] == 'OVRFLW':
                sample_2[i] = sample_1[i]
                outfile.write(sheet+" Sample Overflow: Sample Number: " + str(i + 1) + " absorbance replaced with second measured absorbance " + str(sample_1[i]) + "\n")
        samples = sample_1 + sample_2

    ######defend against overflow ####
        loc = 0
        for item in curve_1:
            if item == "OVRFLW":
                del (curve_1[loc])
                del (curve_2[loc])
                del (curve_yval_1[loc])
                del (curve_yval_2[loc])
                del (dil[loc])
                outfile.write(sheet + " Curve #1 Overflow: Curve Location: " + str(loc +1) + "\n")
                wb[sheet].sheet_properties.tabColor = "1072BA"  
            else:
                loc += 1
        loc = 0
        for item in curve_2:
            if item == "OVRFLW":
                del (curve_1[loc])
                del (curve_2[loc])
                del (curve_yval_1[loc])
                del (curve_yval_2[loc])
                del (dil[loc])
                outfile.write(sheet +" Curve #2 Overflow: Curve Location: " + str(loc +1) + "\n")
                wb[sheet].sheet_properties.tabColor = "1072BA"  
            else:
                loc += 1
              
     #####make sure curve values are in the proper order#####           
        k = 0
        while (k < len(curve_1) - 2):
            if(curve_1[k] < curve_1[k +1]):
                del (curve_2[k])
                del (curve_yval_2[k])
                del (curve_1[k])
                del (curve_yval_1[k])
                del (dil[k])
                outfile.write(sheet+" Curve #1 Descent Issue: Point " + str (k +1) + " was too high and was removed\n")
            else:
                k += 1
                
               
        h = 0
        while (h < len(curve_2) - 2):
            if(curve_2[h]) < curve_2[h + 1]:
               del (curve_2[h])
               del (curve_yval_2[h])
               del (curve_1[h])
               del (curve_yval_1[h])
               del (dil[h])
               outfile.write(sheet+" Curve #2 Descent Issue: Point " + str(h +1) + "was too high and was removed\n")
            else:
               h += 1


    #####Find best combo of points

        curve_length = len(curve_1)
        if curve_length < 4:
            wb[sheet].sheet_properties.tabColor = "1072BA"
            outfile.write(sheet+ " Curve Length Too Short: Data Accuracy May Be Impacted\n")
            
        else:
            best_r = 0
            loc = 0
            best_loc =0
            best_poly = numpy.polyfit((curve_1[0:curve_length] + curve_2[0:curve_length]),(curve_yval_1[0:curve_length] + curve_yval_2[0:curve_length]),3)
            
            #x and y values for curve
            while (loc < curve_length - 4):
                my_x = (curve_1[loc:curve_length] + curve_2[loc:curve_length])
                my_y = curve_yval_1[loc:curve_length] + curve_yval_2[loc:curve_length]

            #####defend samps above curve (20%)####
                above_curve = 0
                for item in samples:
                    #avg of largest amt
                    if item > (my_x[0]):
                        above_curve += 1
                above_curve_per = above_curve/len(samples)
                if above_curve_per > 0.20:
                    go = False

            ###Finding Poly####
                if go == True:
                    poly = numpy.polyfit(my_x, my_y, 3)
                    size = len(my_y)
               ####Find R^2 #######
                    #mean of y
                    mean =  float(sum(my_y)/ size)
                    obt = 0
                    for item in my_y:
                        obt += ((item - mean)**2)
                    #y ^
                    exp= 0
                    for item in my_x:
                        exp +=(((numpy.polyval(poly, item))- mean)**2)
                    r = exp/obt
            ######if new r sqrd is better replace it #####
                    if r > best_r:
                        best_r = r
                        best_loc = loc
                        loc += 1
                    else:
                        loc += 1
                elif go == False:
                    loc += 1

    ###### Averages Using Best Poly get sample amounts #######
    #!!!!!!!!dil factor needs adjustment!!!!!!!!!!!!!
                    #(both its loc and its paste position)
            my_x = curve_1[best_loc:len(curve_1)]+ curve_2[best_loc:len(curve_2)]
            my_y = curve_yval_1[best_loc:len(curve_yval_1)] + curve_yval_2[best_loc:len(curve_yval_2)]
            poly = numpy.polyfit(my_x, my_y, 3)
            spread = True
            if poly[0]< 0 and poly[1]>0 and poly[2] >0 and poly[3] < 0 and max(samples) > max(my_x) and best_loc != 0:
                best_loc -= 1
                my_x = curve_1[best_loc:len(curve_1)]+ curve_2[best_loc:len(curve_2)]
                my_y = curve_yval_1[best_loc:len(curve_yval_1)] + curve_yval_2[best_loc:len(curve_yval_2)]
                best_poly = numpy.polyfit(my_x, my_y, 3)
                spread = False
            elif poly[0]< 0 and poly[1]>0 and poly[2] <0 and poly[3] > 0 and max(samples) > max(my_x)and best_loc != 0:
                best_loc -= 1
                my_x = curve_1[best_loc:len(curve_1)]+ curve_2[best_loc:len(curve_2)]
                my_y = curve_yval_1[best_loc:len(curve_yval_1)] + curve_yval_2[best_loc:len(curve_yval_2)]
                best_poly = numpy.polyfit(my_x, my_y, 3)
                spread = False
            else:
                best_poly = poly

            #Spread of curve points
            
            spread_check = 0
            spread_stdev = statistics.stdev([my_x[0], my_x[int((len(my_x)/2))]])
            spread_avg = (my_x[0] + my_x[int(len(my_x)/2)+1])/2
            if (spread_stdev / spread_avg) > 0.3:
                print ("Bad Spread")
                wb[sheet].sheet_properties.tabColor = "1072BA"
                first_point_keep = my_x[1:len(my_x)]
                first_point_y = my_y[1: len(my_y)]
                second_point_keep = my_x[0 :int(len(my_x)/2)] + my_x[int(len(my_x)/2)+ 1: len(my_x)]
                second_point_y = my_y[0 :int(len(my_y)/2)] + my_y[int(len(my_y)/2)+ 1: len(my_y)]
                first_point_poly = numpy.polyfit(first_point_keep, first_point_y, 3)
                second_point_poly= numpy.polyfit(second_point_keep, second_point_y, 3)
                ####Find R^2 #######
                #point 1
                    #mean of y
                best_r = 0
                mean =  float(sum(first_point_y)/ len(first_point_y))
                obt = 0
                for item in first_point_y:
                    obt += ((item - mean)**2)
                #y ^
                exp= 0
                for item in first_point_keep:
                    exp +=(((numpy.polyval(first_point_poly, item))- mean)**2)
                first_point_r = exp/obt
                #point 2
                best_r = 0
                mean =  float(sum(second_point_y)/ len(second_point_y))
                obt = 0
                for item in second_point_y:
                    obt += ((item - mean)**2)
                #y ^
                exp= 0
                for item in second_point_keep:
                    exp +=(((numpy.polyval(second_point_poly, item))- mean)**2)
                second_point_r = exp/obt
        ######if new r sqrd is better replace it #####
#!!!!!!!!!!!!!!!!!!!!!!!##may want to think more on this!!!!!!!!!!!!!!!!!!!!
                if first_point_r < second_point_r :
                    best_r = first_point_r
                    my_x = first_point_keep
                    my_y = first_point_y
                    best_poly = first_point_poly
                else:
                    best_r = second_point_r
                    my_x = second_point_keep
                    my_y = second_point_y
                    best_poly = second_point_poly
       
            if best_r < 0.7:
                outfile.write(sheet+" R squared value low: " + str(best_r) + "\n")
            
            for item in sample_1:
                samp_amt1.append((numpy.polyval(best_poly, item)*samp_dil))
            for item in sample_2:
                samp_amt2.append((numpy.polyval(best_poly, item)*samp_dil))
            #avg samplse
            #~~~~good~~~~#
            for i in range (0 , 40):
                samp_avg.append((samp_amt1[i] + samp_amt2[i])/2)
            
            #Obt for correction factors
            #~~~~good~~~~#
            for i in range(36,40):
                obtCor.append(samp_avg[i])
            #Expected for correction factors
            #~~~~good~~~~#
            cor = True
            for i in range(row_exp,row_exp + 4):
                if (ws_from.cell(row = i, column = col_exp).value) == None:
                    expCor.append(0)
                    cor = False
                else:
                    expCor.append(ws_from.cell(row = i, column = col_exp).value)
        
            #sum of least squares for best slope
            best_dist = 5
            m = 0
            best_m = 0
            if cor == True:
##################CLEAR THIS ALL UP FOR THE VALUES TOO HIGH OR LOW ####################
                                 
                ##sum of least squares to find the best slope ((y - mx)= d)
                while m < 2.5:
                    sum_dsqrd = 0
                    for i in range (0,3):
                        x = obtCor[i]
                        y = expCor[i]
                        distance_sqrd = (y - (m*x))**2
                        sum_dsqrd += distance_sqrd
                        if sum_dsqrd < best_dist:
                            best_dist = sum_dsqrd
                            best_m = m
                            m += 0.0001
                        else:
                            m += 0.0001
        ###### Paste appropriate Data into Cells #########
            column = 6
            for item in best_poly:
                wb[sheet].cell(row = 16 , column = column).value = item
                column += 1
            #clear the curve cells
            for i in range(col_of_curve_dil,col_of_curve_dil+3):
                for j in range(row_of_curve,row_of_curve + 16):
                    wb[sheet].cell(row = j , column = i).value = ""
            #paste x vals
            row = row_of_curve
            for item in my_x:
                wb[sheet].cell(row = row , column = col_of_curve_x).value = item
                row += 1
            #paste y vals
            row = row_of_curve
            for item in my_y:
                wb[sheet].cell(row = row , column = col_of_curve_y).value = item
                row += 1
            #paste dil vals
            dil = dil[best_loc:len(dil)]
            row = row_of_curve
            for item in dil:
                wb[sheet].cell(row = row , column = 25).value = item
                wb[sheet].cell(row = row + len(dil) , column = col_of_curve_dil).value = item
                row += 1
            #paste amts 
            row = row_samp_start
            for item in samp_amt1:
                wb[sheet].cell(row = row , column = 5).value = item
                row += 1
            row = row_samp_start
            for item in samp_amt2:
                wb[sheet].cell(row = row , column = 9).value = item
                row += 1
                
            #paste control
            wb[sheet].cell(row = 45 , column = 15).value = "SLOPE ="
            if best_m != 0:
                wb[sheet].cell(row = 45 , column = 16).value = best_m
            else:
                wb[sheet].cell(row = 45 , column = 16).value = 1

            #paste obtained RYGB values and their correction
            cor = []
            for i in obtCor:
                if best_m != 0:
                    cor.append(i*best_m)
                else:
                    cor.append(i)
            j = 0
            for i in range(2,6):
                wb[sheet].cell(row = i , column = 22).value = obtCor[j]
                wb[sheet].cell(row = i , column = 23).value = cor[j]
                j+=1
            #clear
            #~~~~good~~~~#
            for i in range(row_samp_start, row_samp_start + 40):
                for j in range(11,14):
                    wb[sheet].cell(row = i , column = j).value = None
            #paste avg (11)
            #~~~~good~~~~#
            row = row_samp_start
            for item in samp_avg:
                wb[sheet].cell(row = row , column = 11).value = item
                row += 1
         
            #calculate and paste stdev (12)
            #~~~~good~~~~#
            for i in range (0,40):
                std = statistics.stdev([samp_amt1[i],samp_amt2[i]])
                stdv.append(std)
            for i in range (0 , 40):
                wb[sheet].cell(row = i + row_samp_start, column = 12).value = stdv[i]
            #calculate and paste CV (13)
            #~~~~good~~~~#
            for i in range (0,40):
                cv.append(stdv[i]/samp_avg[i])
            for i in range (0 , 40):
                wb[sheet].cell(row = i + row_samp_start, column = 13).value = cv[i]

                
        ###### Make Graphs ######       
            
            CURVE_VS_SAMP(sheet, my_x, samples )
            CURVE (sheet, my_x, my_y, best_poly)
            if max(expCor) != 0:
                EXP_VS_OBT (sheet, obtCor, expCor)
                if (best_m) > 1.5:
                    outfile.write(sheet + " : Obtained Values From Standards Are Much Higher Than Expected (m value): " + str(best_m) + "\n")
                elif (best_m)< 0.5:
                    outfile.write(sheet + " : Obtained Values From Standards Are Much Lower Than Expected (m value): " + str(best_m) + "\n")
            BAR_EXP_VS_OBT (sheet, obtCor, expCor , cor)
            AVERAGES (sheet, samp_avg)
    

    ###### Save corrected file ####

    os.chdir(path2)
    wb.save("Optimized - " + infile_name)
    outfile.write("\n\n\n")
    outfile.close()
##    the_go = input(str("Would you like to Optimize another File? (Enter 'N' to Escape): "))
##    if the_go == 'N' or the_go == 'n':
##        Good_to_Go = False

###### Future Projects could use this program very simply
        #pros and cons of template
        #how to make program more user friendly
