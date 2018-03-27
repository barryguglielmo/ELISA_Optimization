##import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import os
os.chdir('C:/Users/bag019/Desktop')
wb = load_workbook('Database Fo Real.xlsx',data_only=True)
ws_from = wb.get_sheet_by_name('WP AI')
##data = 'Database Fo Real.csv'
##x = pd.read_csv(data, header = 0, usecols = ['S.ID'])
##y = pd.read_csv(data,header = 0, usecols = ['Average 1'], index_col = ['Average 1'])
##
my_list_y = []
my_list_x = []
for j in range(3,1837):
   data_from1 = ws_from.cell(row = j, column = 5).value
   my_list_x.append(data_from1)
for j in range(3,1837):
   data_from2 = ws_from.cell(row = j, column = 2).value
   my_list_y.append(data_from2)
ind = np.arange[len(my_list_x)]
plt.bar(ind, my_list_y)
plt.show()
##print (my_list_x)
##print (my_list_y)
##plt.bar(my_list_x,my_list_y,label = 'Yah doode')
##plt.show()

